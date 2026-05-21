import logging
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.check_result import ComplianceReport, DocumentCheckResult, STRCheckResult

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Colours
# ---------------------------------------------------------------------------

_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_ORANGE = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
_GREY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_BLUE_HEADER = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_ROW_ALT = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

_STATUS_FILL = {
    "pass": _GREEN,
    "fail": _RED,
    "partial": _ORANGE,
    "not_applicable": _GREY,
    "error": _RED,
}

_STATUS_LABEL = {
    "pass": "ATITINKA",
    "fail": "NEATITINKA",
    "partial": "DALINAI",
    "not_applicable": "NETAIKOMA",
    "error": "KLAIDA",
}

_CATEGORY_LABEL = {
    "str": "STR",
    "law": "Įstatymas",
    "hn": "HN norma",
    "other": "Kita",
}

_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _header_font(white: bool = True) -> Font:
    return Font(bold=True, color="FFFFFF" if white else "1F4E79", size=11)


def _set_col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _style_header_row(ws, row: int, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _BLUE_HEADER
        cell.font = _header_font()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER


def _apply_row_style(cell, alt: bool = False) -> None:
    if not cell.fill or cell.fill.fill_type == "none":
        if alt:
            cell.fill = _ROW_ALT
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = _BORDER


# ---------------------------------------------------------------------------
# Sheet 1: STR Compliance
# ---------------------------------------------------------------------------

_STR_HEADERS = [
    "Kategorija",
    "Kodas",
    "Pavadinimas",
    "Minimas\ndokumente",
    "Atitikties\nbūklė",
    "Komentaras",
    "Šaltinis\n(straipsnis)",
    "Puslapiai",
    "Patikimumas",
]
_STR_WIDTHS = [12, 22, 52, 12, 14, 60, 22, 12, 14]


def _write_str_sheet(ws, results: list[STRCheckResult]) -> None:
    ws.title = "STR atitiktis"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 36

    for col, header in enumerate(_STR_HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, 1, len(_STR_HEADERS))
    _set_col_widths(ws, _STR_WIDTHS)

    for row_idx, result in enumerate(results, start=2):
        alt = row_idx % 2 == 0
        status_fill = _STATUS_FILL.get(result.status, _GREY)
        pages_str = ", ".join(str(p) for p in sorted(result.page_references)) or "—"

        values = [
            _CATEGORY_LABEL.get(result.category, result.category),
            result.code,
            result.full_name,
            "Taip" if result.present_in_doc else "Ne",
            _STATUS_LABEL.get(result.status, result.status),
            result.comment,
            result.citation or "—",
            pages_str,
            result.confidence.upper() if result.confidence else "—",
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _apply_row_style(cell, alt)
            # Colour the status cell
            if col_idx == 5:
                cell.fill = status_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="top")


# ---------------------------------------------------------------------------
# Sheet 2: Mandatory Documents
# ---------------------------------------------------------------------------

_DOC_HEADERS = ["Privalomasis dokumentas", "Rastas", "Komentaras"]
_DOC_WIDTHS = [44, 10, 70]


def _write_docs_sheet(ws, results: list[DocumentCheckResult]) -> None:
    ws.title = "Privalomieji dokumentai"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    for col, header in enumerate(_DOC_HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)
    _style_header_row(ws, 1, len(_DOC_HEADERS))
    _set_col_widths(ws, _DOC_WIDTHS)

    for row_idx, result in enumerate(results, start=2):
        alt = row_idx % 2 == 0
        found_label = "Taip" if result.found else "Ne"
        found_fill = _GREEN if result.found else _RED

        values = [result.document_name, found_label, result.comment]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            _apply_row_style(cell, alt)
            if col_idx == 2:
                cell.fill = found_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="top")


# ---------------------------------------------------------------------------
# Sheet 3: Summary
# ---------------------------------------------------------------------------


def _write_summary_sheet(ws, report: ComplianceReport) -> None:
    ws.title = "Suvestinė"
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 20

    def row(label: str, value) -> None:
        r = ws.max_row + 1
        a = ws.cell(row=r, column=1, value=label)
        b = ws.cell(row=r, column=2, value=value)
        a.font = Font(bold=True)
        a.border = _BORDER
        b.border = _BORDER
        b.alignment = Alignment(horizontal="left")

    ws.cell(row=1, column=1, value="TIKRINIMO SUVESTINĖ")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="1F4E79")
    ws.merge_cells("A1:B1")

    ws.append([])
    row("Dokumentas", report.filename)
    row("Tikrinimo data", report.checked_at.strftime("%Y-%m-%d %H:%M"))
    row("Puslapių skaičius", report.total_pages)

    ws.append([])
    row("Patikrinta normatyvų", len(report.str_results))
    row("✓ Atitinka", report.pass_count)
    row("✗ Neatitinka", report.fail_count)
    row("~ Dalinai atitinka", report.partial_count)
    row(
        "— Netaikoma / neklaida",
        len(report.str_results) - report.pass_count - report.fail_count - report.partial_count,
    )

    ws.append([])
    row("Privalomieji dokumentai (iš viso)", len(report.document_results))
    row("Rasti dokumentai", report.docs_found_count)
    row("Nerasti dokumentai", len(report.document_results) - report.docs_found_count)

    if report.warnings:
        ws.append([])
        ws.cell(row=ws.max_row + 1, column=1, value="Įspėjimai").font = Font(
            bold=True, color="FF0000"
        )
        for warning in report.warnings:
            ws.append(["", warning])


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


class ExcelReporter:
    def generate(self, report: ComplianceReport) -> bytes:
        """Generate an Excel workbook and return it as bytes."""
        logger.info("Generating Excel report for: %s", report.filename)

        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        ws_str = wb.create_sheet()
        _write_str_sheet(ws_str, report.str_results)

        ws_docs = wb.create_sheet()
        _write_docs_sheet(ws_docs, report.document_results)

        ws_summary = wb.create_sheet()
        _write_summary_sheet(ws_summary, report)

        # Make summary first
        wb.move_sheet(ws_summary, offset=-wb.index(ws_summary))

        buffer = BytesIO()
        wb.save(buffer)
        logger.info("Excel report generated (%d bytes)", buffer.tell())
        return buffer.getvalue()
